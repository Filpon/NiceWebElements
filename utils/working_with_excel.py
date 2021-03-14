import openpyxl
import os

from datetime import datetime


class ExcelWorksWithSite:
    """
	Class for working with Site via Excel
	"""

    def __init__(self, step=0):
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.step = step
        self.second_step_row_start = 0

    def increment_self_step(self):
        """
        Incrementing step number
        """
        self.step += 1

    def input_data_generator(self, input_data_list: list):
        """
        Generator for input data to row
        :param input_data_list: List with data for generator
        :return: Bool function result
        """
        try:
            for row in input_data_list:
                if not isinstance(row, list):
                    yield row
                else:
                    if row != input_data_list[0]:
                        for new_row in row:
                            yield new_row
            return True
        except Exception as excp:
            print(excp)
            return False

    def paste_headers(self, headers: list):
        """
        Pasting headers
        :param headers: Headers for excel
        """
        generator_object_headers = self.input_data_generator(input_data_list=headers)
        for cell_column in range(1, len(headers) + 1):
            row_to_paste = next(generator_object_headers)
            if not row_to_paste:
                break
            self.worksheet.cell(row=1, column=cell_column).value = row_to_paste

    def write_to_excel_file(self, list_of_data: list, stop_writing_data=False):
        """
        Writing data to excel files
        :param list_of_data:
        :param stop_writing_data:
        """
        generator_object_rows = self.input_data_generator(input_data_list=list_of_data)
        if self.second_step_row_start == 0:
            self.second_step_row_start = len(list_of_data) + 1
        range_row = range(2, len(list_of_data) + 1)
        range_column = range(1, 4) if self.step == 0 else range(4, 7)
        for cell_row in range_row:
            for cell_column in range_column:
                try:
                    row_to_paste = next(generator_object_rows)
                except StopIteration:
                    row_to_paste = False
                    stop_writing_data = True
                if not row_to_paste:
                    break
                self.worksheet.cell(row=cell_row, column=cell_column).value = (
                    float(row_to_paste.replace(",", "."))
                    if cell_column not in [1, 4]
                    else row_to_paste
                )
            if stop_writing_data:
                break

    def save_file(self, file_name: str):
        """
        Output file saving
        :param file_name: Name of output file
        """
        current_filename = os.path.join(
            os.getcwd(),
            f"{file_name.replace('.xlsx', '')}_{datetime.now().strftime('%d-%m-%Y %H-%M-%S')}.xlsx",
        )
        try:
            self.workbook.save(current_filename)
            return current_filename
        except Exception:
            return None
